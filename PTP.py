import time

import  openpyxl
import pandas as  pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

print('**********请稍等,正在匹配PTP list和DevMgrInfo**********')
border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))#加线框
font=Font(size=12,name='Calibri')#设置样式
align = Alignment(horizontal='center', vertical='center',wrap_text=True)#设置对齐方式
#align_1=Alignment(wrap_text=True)#设置单元格内容自动换行显示
ptp_dict={}
mgr_dict={}
ptpdrivername_list=[]
ptpdriverversion_list=[]
mgrdrivername_list=[]
mgrdriverversion_list=[]
wb=openpyxl.Workbook()#创建新的excel
sheet=wb['Sheet']
sheet.column_dimensions['A'].width = 40 #把ABDEF列的单元格拉长延申一些
sheet.column_dimensions['B'].width = 68
sheet.column_dimensions['D'].width = 50
sheet.column_dimensions['E'].width = 22

target_cell=sheet.cell(column=1, row=1)#设置生成结果的表头
target_cell.value='PTP Driver Name'
target_cell.font=font
target_cell.alignment=align
target_cell.border = border

target_cell=sheet.cell(column=2, row=1)
target_cell.value='PTP Driver Version'
target_cell.font=font
target_cell.alignment=align
target_cell.border = border

target_cell=sheet.cell(column=3, row=1)
target_cell.value='VS'
target_cell.font=font
target_cell.alignment=align
target_cell.border = border

target_cell=sheet.cell(column=4, row=1)
target_cell.value='Mgr Driver Name'
target_cell.font=font
target_cell.alignment=align
target_cell.border = border

target_cell=sheet.cell(column=5, row=1)
target_cell.value='Mgr Driver Version'
target_cell.font=font
target_cell.alignment=align
target_cell.border = border

ptp=pd.read_excel(r'C:\ptp.xlsx',header=6)
ptpdriver_name=ptp['Driver/Utility Name']
ptpdriver_version=ptp['VersionNo']

mgr=pd.read_excel(r'C:\DevMgrInfo.xlsx')
mgrdriver_name=mgr['Device']
mgrdriver_version=mgr['DriverVersion']

row_number=2

for name in ptpdriver_name:
    ptpdrivername_list.append(name)
for version in ptpdriver_version:
    ptpdriverversion_list.append(version)
for i in range(0,len(ptpdriverversion_list)):
    target_cell=sheet.cell(column=1, row=row_number)#因为列数是不变的
    target_cell.value = ptpdriver_name[i]
    target_cell.font = font
    target_cell.alignment = align
    target_cell.border = border
    target_cell=sheet.cell(column=2, row=row_number)
    target_cell.value = ptpdriver_version[i]
    target_cell.font = font
    target_cell.alignment = align
    target_cell.border = border
    target_cell=sheet.cell(column=3, row=row_number)
    target_cell.value ='VS'
    target_cell.font = font
    target_cell.alignment = align
    target_cell.border = border
    ptp_dict[ptpdrivername_list[i]]=ptpdriverversion_list[i]# 创建{'Lenovo Intelligent Thermal Solution Driver': '2.1.14.0'}样式,为了后面结果匹配做工作
    row_number=row_number+1

for mgrname in mgrdriver_name:
    mgrdrivername_list.append(mgrname)
for mgrdriverversion in mgrdriver_version:
    mgrdriverversion_list.append(mgrdriverversion)###这部分把mgrlist里的driver name 和 version组成字典
for i in range (0,len(mgrdriverversion_list)):
    mgr_dict[mgrdrivername_list[i]]=mgrdriverversion_list[i]

result_row=2
#整体逻辑如下
for x in ptpdriverversion_list:
    SUT_Driver_names=[key for key,val in mgr_dict.items() if val==x] #找到mgr 字典里有没有一样driver版本的。有可能是一个driver 版本对应多个driver name
    if  len(SUT_Driver_names) ==1: #说明一个driver对应一个name
        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
        target_cell.value = x
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill=openpyxl.styles.PatternFill(start_color='00FF00',end_color='00FF00',fill_type='solid')#填充绿色背景颜色
        target_cell.border = border
        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
        target_cell.value = SUT_Driver_names[0]
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
        target_cell.border = border
    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
        target_cell.value = x
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        target_cell.border = border
        for i in range(0,len(SUT_Driver_names)):
            target_cell = sheet.cell(column=4, row=result_row)  #
            original_content = target_cell.value#获取本来的单元格内容，方便后面追究
            if original_content==None:
                original_content=' '# 如果是none ，后面none+字符串会包错，所以随便定义一个字符
            target_cell.value = original_content+SUT_Driver_names[i]+'\n'
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        target_cell.border = border
    elif len(SUT_Driver_names) ==0 and 'INF' not in x and '(' not in x and ':' not in x and '\n' not in x and ',' not in x:# 说明真的没找到值且driver 版本也是正常的模式10.1.57.4这种，说明真没有匹配的
        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
        target_cell.value = 'NA'
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
        target_cell.border = border
        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
        target_cell.value = 'NA'
        target_cell.font = font
        target_cell.alignment = align
        target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
        target_cell.border = border
################后面的情况比较复杂，后续如果实在变化多端，就把PTP整理成前面以上代码的3种情况就完全可以跑tool了，每一行就是纯driver1.0.16.3这种

    else:#因为driver格式问题没找到对应的，这里又分为了几种情况如下
        if "(Component INF)" in x and ',' in x:  #情况格式一：1.0.2.0 (Extension INF), 10.1.10.0 (Component INF)，这里考虑只有一个driver只有一个(Component INF)，如果多个代码要变一下，需要再版本和name栏位都要追加
            list=x.split(',')#按照逗号把字符串转列表
            for any in list:
                if "(Component INF)" in any:#Extension INF不需要匹配所以直接匹配Component INF
                    newstr=any.replace("(Component INF)","")#删掉(Component INF)字眼
                    newstr_1=newstr.replace(" ","")#再去掉空格，留下10.1.10.0字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]#这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        target_cell.value = newstr_1
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
                        target_cell.value = SUT_Driver_names[0]
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                        target_cell.border = border
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        target_cell.value = newstr_1
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        target_cell.border = border
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + '\n'
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        target_cell.border = border
                    else:#标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        target_cell.value = 'NA'
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
                        target_cell.border = border
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
                        target_cell.value = 'NA'
                        target_cell.font = font
                        target_cell.alignment = align
                        target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
                        target_cell.border = border
        elif "(Component INF)" in x and ','  not in x and '\n ' not in x:   #情况2：只有1个(Component INF)，        4.3.11.0 (Component INF)
            newstr = x.replace("(Component INF)", "")  # 删掉(Component INF)字眼
            newstr_1 = newstr.replace(" ", "")  # 再去掉空格，留下10.1.10.0字眼
            SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
            if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.value = newstr_1
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
                target_cell.value = SUT_Driver_names[0]
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
            elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                target_cell.value = newstr_1
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                target_cell.border = border
                for i in range(0, len(SUT_Driver_names)):
                    target_cell = sheet.cell(column=4, row=result_row)  #
                    original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                    if original_content == None:
                        original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                    target_cell.value = original_content + SUT_Driver_names[i] + '\n'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                target_cell.border = border
            else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.value = 'NA'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头
                target_cell.value = 'NA'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色背景颜色
                target_cell.border = border
        elif "(Senary Audio)" in x  and "(Intel ISST)" in x and "," in x:#3.48.42.4(Senary Audio), 20.40.11433.0(Intel ISST)这种情况. 这种必须要driver name 和版本栏位都要追加值,且都要找值
            list = x.split(',')  # 按照逗号把字符串转列表
            all_list=[]##用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
            for any in list:
                if "(Senary Audio)" in any:
                    newstr = any.replace("(Senary Audio)", "")  # 删掉((Senary Audio)字眼
                    newstr_1 = newstr.replace(" ", "")  # 再去掉空格，留下3.48.42.4字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass'+'\n'#driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0]+":" + newstr_1+'\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 +  ':Pass'+'\n'#driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] +":"+newstr_1+'\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                else:
                    newstr = any.replace("(Intel ISST)", "")  # 删掉(Component INF)字眼
                    newstr_1 = newstr.replace(" ", "")  # 再去掉空格，留下3.48.42.4字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass'+'\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 +':Pass'+ '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':' + 'NA' + "\n"
            list1=[]# 收集空列表
            list2=[]#收集非空列表
            for i in all_list:
                if i ==[]:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1==[] :#代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
            elif list2==[] :#代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.value = 'NA'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
            elif list1 !=[] and list2 !=[]: # 代表都有收集到，设置为黄色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
        elif 'realtek' in x  and 'Sunplus' in x and 'Sonix' in x:#realtek:10.0.22000.20346 Sunplus:5.0.18.242   Sonix:10.13.22621.30这种情况
            list=x.split(' ')
            all_list=[]
            for any in list:
                if "realtek:" in any:
                    newstr_1 = any.replace("realtek:", "")  # 删掉realtek:字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Sunplus:" in any:
                    newstr_1 = any.replace("Sunplus:", "")  # 删掉Sunplus:字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':' + 'NA' + "\n"
                elif "Sonix:" in any:
                    newstr_1 = any.replace("Sonix:", "")  # 删掉Sunplus:字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':' + 'NA' + "\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
        elif "Driver Package Version" in x and "\n" in x and "Quectel GNSS Sensor Device" in x and "Quectel MBIHV Mobile Broadband Firmware Device" in x and "Quectel USB Composition device" in x and "Quectel USB Device for Legacy Serial Communication" in x: #'1.0.0.15 (Driver Package Version)\n                    1.0.2.11 (Quectel GNSS Sensor Device)\n                    5.0.0.26 (Quectel MBIHV Mobile Broadband Firmware Device)\n
            list=x.split('\n')
            all_list=[]
            for any in list:
                if "Driver Package Version" in any:#Driver Package Version"不需要对照
                    pass
                elif "Quectel GNSS Sensor Device" in any:
                    newstr=any.replace("(Quectel GNSS Sensor Device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Quectel MBIHV Mobile Broadband Firmware Device" in any:
                    newstr=any.replace("(Quectel MBIHV Mobile Broadband Firmware Device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Quectel USB Composition device" in any:
                    newstr=any.replace("(Quectel USB Composition device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Quectel USB Device for Legacy Serial Communication" in any:
                    newstr=any.replace("(Quectel USB Device for Legacy Serial Communication)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.J
                target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色色色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
        elif "Driver Package Version" in x and "\n" in x and "NetPrisma(R) Location" in x and "NetPrisma MBIHV Mobile Broadband Firmware Device" in x and "NetPrisma USB Composition device" in x: #'1.0.0.15 (Driver Package Version)\n                    1.0.2.11 (Quectel GNSS Sensor Device)\n                    5.0.0.26 (Quectel MBIHV Mobile Broadband Firmware Device)\n
            list=x.split('\n')
            all_list=[]
            for any in list:
                if "Driver Package Version" in any:#Driver Package Version"不需要对照
                    pass
                elif "(NetPrisma(R) Location)" in any:
                    newstr=any.replace("(NetPrisma(R) Location)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "(NetPrisma MBIHV Mobile Broadband Firmware Device)" in any:
                    newstr=any.replace("(NetPrisma MBIHV Mobile Broadband Firmware Device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "(NetPrisma USB Composition device)" in any:
                    newstr=any.replace("(NetPrisma USB Composition device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.J
                target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色色色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
        elif "Driver Package Version" in x and "\n" in x and "NetPrisma(R) Location" in x and "NetPrisma MBIHV Mobile Broadband Firmware Device" in x and "HighSpeed USB Composite Device(UDE)" in x: #'1.0.0.15 (Driver Package Version)\n                    1.0.2.11 (Quectel GNSS Sensor Device)\n                    5.0.0.26 (Quectel MBIHV Mobile Broadband Firmware Device)\n
            list=x.split('\n')
            all_list=[]
            for any in list:
                if "Driver Package Version" in any:#Driver Package Version"不需要对照
                    pass
                elif "(NetPrisma(R) Location)" in any:
                    newstr=any.replace("(NetPrisma(R) Location)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "(NetPrisma MBIHV Mobile Broadband Firmware Device)" in any:
                    newstr=any.replace("(NetPrisma MBIHV Mobile Broadband Firmware Device)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "HighSpeed USB Composite Device(UDE)" in any:
                    newstr=any.replace("HighSpeed USB Composite Device(UDE)", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色 target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色色色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
        elif "(Synaptics)" in x  and "(Goodix)" in x and " " in x:#6.0.51.1136(Synaptics) 3.4.51.1030(Goodix)
            list = x.split(' ')  # 按照逗号把字符串转列表
            all_list=[]##用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
            for any in list:
                if "(Synaptics)" in any:
                    newstr_1 = any.replace("(Synaptics)", "")  # 删掉((Senary Audio)字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass'+'\n'#driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0]+":" + newstr_1+'\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 +  ':Pass'+'\n'#driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] +":"+newstr_1+'\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                else:
                    newstr_1 = any.replace("(Goodix)", "")  # 删掉(Component INF)字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass'+'\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 +':Pass'+ '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':' + 'NA' + "\n"
            list1=[]# 收集空列表
            list2=[]#收集非空列表
            for i in all_list:
                if i ==[]:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1==[] :#代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
            elif list2==[] :#代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.value = 'NA'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
            elif list1 !=[] and list2 !=[]: # 代表都有收集到，设置为黄色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
        elif "(Base driver)" in x  and "(Component Extension)" in x and " " in x:#2.2.10201.2(Base driver)                   9.0.11900.49842(Component Extension)
            list = x.split('          ')  # 按照逗号把字符串转列表
            all_list=[]##用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
            for any in list:
                if "(Component Extension)" in any:
                    pass
                elif "(Base driver)" in any:
                    newstr_1 = any.replace("(Base driver)", "")  # 删掉(Component INF)字眼
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)
                    if len(SUT_Driver_names) == 1:  # 说明一个driver对应一个name
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass'+'\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 +':Pass'+ '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':' + 'NA' + "\n"
            list1=[]# 收集空列表
            list2=[]#收集非空列表
            for i in all_list:
                if i ==[]:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1==[] :#代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
            elif list2==[] :#代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.value = 'NA'
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
            elif list1 !=[] and list2 !=[]: # 代表都有收集到，设置为黄色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
        elif "MEI" in x and "\n" in x and "SOL" in x and "DAL" in x and "LMS" in x and 'WMI' in x and 'Wiman' in x:#'MEI: 2428.6.2.0\n  SOL: 2405.6.0.0\n  DAL: 1.46.2024.0221\n  LMS: 2430.6.13.0\n  WMI: 2428.6.4.0\n  Wiman: 2433.90.33.0'
            list=x.split('\n')
            all_list=[]
            for any in list:
                if "MEI" in any:
                    newstr=any.replace("MEI:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "SOL" in any:
                    newstr=any.replace("SOL:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "DAL" in any:
                    newstr=any.replace("DAL:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "LMS" in any:
                    newstr=any.replace("LMS:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "WMI" in any:
                    newstr=any.replace("WMI:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Wiman" in any:
                    newstr=any.replace("Wiman:", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色 target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色色色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
        elif "ELAN" in x and "\n" in x and "TrackPoint Device" in x and "TrackPoint Componen" in x and "ETD HSA Device" in x and 'Trackpoint for Thinkpad' in x:#(Driver)\n  ELAN\u3000TrackPoint Device  31.21.45.1\n  Elan TrackPoint Component 31.21.0.6\n  ETD HSA Device                 31.21.0.5\n   (App)\n               ELAN Trackpoint for Thinkpad  24.121.52.0
            list=x.split('\n')
            all_list=[]
            for any in list:
                if "TrackPoint Device" in any:
                    newstr=any.replace("ELAN\u3000TrackPoint Device", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "TrackPoint Component" in any:
                    newstr=any.replace("Elan TrackPoint Component", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "ETD HSA Device" in any:
                    newstr=any.replace("ETD HSA Device", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
                elif "Trackpoint for Thinkpad" in any:
                    newstr=any.replace("ELAN Trackpoint for Thinkpad", "")
                    newstr_1 = newstr.replace(" ", "")  #
                    SUT_Driver_names = [key for key, val in mgr_dict.items() if val == newstr_1]  # 这种标准格式的话列表长度要么就是0要么就是1要么就是大于1了
                    all_list.append(SUT_Driver_names)  ###用来判断最终单元格的颜色，找到其中一个driver就是警告黄色，全找到就是绿色，都没找到就是灰色
                    if len(SUT_Driver_names) == 1:
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                        original_content = target_cell.value
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + SUT_Driver_names[0] + ":" + newstr_1 + '\n'
                    elif len(SUT_Driver_names) > 1:  # 说明一个driver对应多个name,需要在单元格追加内容
                        target_cell = sheet.cell(column=5, row=result_row)  # driver 版本是唯一的一个不用追加
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content + newstr_1 + ':Pass' + '\n'  # driver 版本也要追加
                        for i in range(0, len(SUT_Driver_names)):
                            target_cell = sheet.cell(column=4, row=result_row)  #
                            original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                            if original_content == None:
                                original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                            target_cell.value = original_content + SUT_Driver_names[i] + ":" + newstr_1 + '\n'
                    else:  # 标准driver格式却没找到对应的就是真没有,NA掉
                        target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                        original_content = target_cell.value  # 获取本来的单元格内容，方便后面追究
                        if original_content == None:
                            original_content = ' '  # 如果是none ，后面none+字符串会包错，所以随便定义一个字符
                        target_cell.value = original_content+newstr_1+':'+ 'NA'+"\n"
            list1 = []  # 收集空列表
            list2 = []  # 收集非空列表
            for i in all_list:
                if i == []:
                    list1.append(i)
                else:
                    list2.append(i)
            if list1 == []:  # 代表没有空列表，设置绿色背景
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 填充绿色背景颜色 target_cell.border = border
            elif list2 == []:  # 代表没有收集到非空列表，所以都是空列表，设置为灰色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')  # 填充灰色色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.value = 'NA'
                target_cell.fill = openpyxl.styles.PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                target_cell.border = border
            elif list1 != [] and list2 != []:  # 代表都有收集到，设置为黄色色色
                target_cell = sheet.cell(column=5, row=result_row)  # 设置生成结果的表头
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border
                target_cell = sheet.cell(column=4, row=result_row)  # 设置生成结果的表头，driver name也要追加
                target_cell.font = font
                target_cell.alignment = align
                target_cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 填充黄色背景颜色
                target_cell.border = border





    result_row = result_row + 1
wb.save(r'C:\result.xlsx')
print('**********已完成,结果请参考C盘根目录result.xlsx**********')
time.sleep(3)


